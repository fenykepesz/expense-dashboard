import sys
from pathlib import Path

import openpyxl
import pytest

sys.path.insert(0, str(Path(__file__).parent.parent))
sys.path.insert(0, str(Path(__file__).parent.parent / "tools"))

from holdings_filing_to_json import parse_holdings_filing


def _build_cover_sheet(wb, institution_name="Test Insurance Ltd", institution_reg_number="123456789",
                        period_year=2026, period_quarter="01", shifted=False):
    ws = wb.active
    ws.title = "עמוד פתיחה"
    offset = 1 if shifted else 0  # simulates a real-world layout where the
    # label lands one column further right than usual (confirmed to happen
    # between real institutions) — offset=True exercises that resilience.
    rows = [
        ("יש לבחור את רבעון הדיווח:", period_quarter),
        ("יש לבחור את שנת הדיווח:", period_year),
        ("יש לבחור את הגוף המוסדי:", institution_name),
        ("ח.פ. הגוף המוסדי:", institution_reg_number),
    ]
    for i, (label, value) in enumerate(rows, start=1):
        ws.cell(row=i, column=1 + offset, value=label)
        ws.cell(row=i, column=4 + offset, value=value)
    return ws


def _build_stocks_sheet(wb, rows):
    """rows: list of (kupa_number, track_number, issuer_name, issuer_number,
    security_name, security_number, pct_of_track, country, sector, currency,
    fair_value_thousands). fair_value_thousands is in thousands ILS, matching
    the real file's convention — the parser multiplies by 1000 on read."""
    ws = wb.create_sheet("מניות מבכ ויהש")
    headers = [
        "מספר קופה/קרן/ח.פ.", "מספר מסלול", "שם מנפיק", "מספר מנפיק",
        "שם נייר ערך", "מספר נייר ערך", "מדינה לפי חשיפה כלכלית",
        "ענף מסחר", "מטבע פעילות", "שיעור מנכסי אפיק ההשקעה",
        "שווי הוגן (באלפי ש\"ח)",
    ]
    ws.append(headers)
    for r in rows:
        (kupa, track, issuer_name, issuer_number, sec_name, sec_number,
         country, sector, currency, pct, fair_value_thousands) = r
        ws.append([kupa, track, issuer_name, issuer_number, sec_name, sec_number,
                   country, sector, currency, pct, fair_value_thousands])
    return ws


def _save(wb, tmp_path, name="filing.xlsx"):
    path = tmp_path / name
    wb.save(path)
    return path


def test_parses_cover_metadata_and_matches_rows(tmp_path):
    wb = openpyxl.Workbook()
    _build_cover_sheet(wb, institution_reg_number="123456789", period_year=2026, period_quarter="01")
    _build_stocks_sheet(wb, [
        ("123456789", "5", "Acme Corp", "999", "Acme Ord", "IL0001", "Israel", "Tech", "ILS", 0.25, 100),
        ("123456789", "9", "Other Corp", "888", "Other Ord", "IL0002", "Israel", "Tech", "ILS", 0.10, 50),
    ])
    path = _save(wb, tmp_path)

    funds = [{"id": 1, "institution_reg_number": "123456789", "track_number": "5"}]
    result = parse_holdings_filing(path, funds)

    assert result["institution_reg_number"] == "123456789"
    assert result["period_year"] == 2026
    assert result["period_quarter"] == 1
    assert len(result["rows"]) == 1
    assert result["rows"][0]["fund_id"] == 1
    assert result["rows"][0]["security_name"] == "Acme Ord"
    assert result["rows"][0]["pct_of_track"] == 0.25
    assert result["rows"][0]["fair_value_ils"] == pytest.approx(100_000)  # 100 thousand ILS -> real ILS
    assert result["unmatched_track_count"] == 1  # the track-9 row


def test_cover_sheet_label_shifted_by_one_column_still_resolves(tmp_path):
    """Confirmed real-world case: one institution's cover sheet has an extra
    leading marker column, shifting the label one column right of another
    institution's layout."""
    wb = openpyxl.Workbook()
    _build_cover_sheet(wb, institution_reg_number="555", period_year=2025, period_quarter="04", shifted=True)
    _build_stocks_sheet(wb, [("555", "1", "X", "1", "X Ord", "IL9999", "", "", "ILS", 0.5, 10)])
    path = _save(wb, tmp_path)

    funds = [{"id": 1, "institution_reg_number": "555", "track_number": "1"}]
    result = parse_holdings_filing(path, funds)
    assert result["institution_reg_number"] == "555"
    assert result["period_year"] == 2025
    assert result["period_quarter"] == 4
    assert len(result["rows"]) == 1


def test_missing_cover_sheet_raises(tmp_path):
    wb = openpyxl.Workbook()
    wb.active.title = "Some Other Sheet"
    path = _save(wb, tmp_path)
    with pytest.raises(ValueError):
        parse_holdings_filing(path, [])


def test_no_recognized_holdings_sheets_raises(tmp_path):
    wb = openpyxl.Workbook()
    _build_cover_sheet(wb)
    ws = wb.create_sheet("Unrelated Sheet")
    ws.append(["foo", "bar"])
    ws.append([1, 2])
    path = _save(wb, tmp_path)
    with pytest.raises(ValueError):
        parse_holdings_filing(path, [])


def test_unmapped_but_holdings_shaped_sheet_imports_as_other(tmp_path):
    wb = openpyxl.Workbook()
    _build_cover_sheet(wb, institution_reg_number="1")
    ws = wb.create_sheet("A Brand New Sheet Type")
    ws.append(["מספר מסלול", "שם מנפיק", "שם נייר ערך", "שיעור מנכסי אפיק ההשקעה", "שווי הוגן (באלפי ש\"ח)"])
    ws.append(["1", "Some Issuer", "Some Security", 0.5, 20])
    path = _save(wb, tmp_path)

    funds = [{"id": 1, "institution_reg_number": "1", "track_number": "1"}]
    result = parse_holdings_filing(path, funds)
    assert len(result["rows"]) == 1
    assert result["rows"][0]["instrument_type"] == "other"
    assert result["rows"][0]["fair_value_ils"] == pytest.approx(20_000)
    assert any(u["sheet_name"] == "A Brand New Sheet Type" for u in result["unrecognized_sheets"])


def test_other_derivatives_sheet_classified_per_row_by_asset_type(tmp_path):
    """Confirmed real-world case: לא סחיר נגזרים אחרים bundles several
    genuinely different OTC derivative types (FX, interest rate, equity,
    inflation) under one sheet name — a flat sheet-name mapping can't tell
    them apart, but the sheet's own סוג הנכס (Asset Type) column can."""
    from holdings_filing_to_json import OTHER_DERIVATIVES_SHEET_NAME

    wb = openpyxl.Workbook()
    _build_cover_sheet(wb, institution_reg_number="1")
    ws = wb.create_sheet(OTHER_DERIVATIVES_SHEET_NAME)
    ws.append(["מספר מסלול", "צד נגדי - Counterparty", "טיקר", "סוג הנכס",
               "שיעור מנכסי אפיק ההשקעה", "שווי הוגן (באלפי ש\"ח)"])
    ws.append(["1", "CITIUS33", "USDILS", 'מט"ח', 0.1, 10])
    ws.append(["1", "POALILIT", "M1IN INDEX", "מניות לרבות מדדי מניות", 0.1, 5])
    ws.append(["1", "LUMIILIT", "IRS-1", "ריבית ואג\"ח", 0.1, 3])
    ws.append(["1", "GSILGB2X", "CPI-1", "מדד המחירים לצרכן", 0.1, 1])
    ws.append(["1", "UNKNOWN33", "MYSTERY-1", "משהו לא מוכר", 0.1, 2])  # unmapped asset type
    path = _save(wb, tmp_path)

    funds = [{"id": 1, "institution_reg_number": "1", "track_number": "1"}]
    result = parse_holdings_filing(path, funds)
    by_ticker = {r["security_name"]: r["instrument_type"] for r in result["rows"]}
    assert by_ticker["USDILS"] == "fx_swap"
    assert by_ticker["M1IN INDEX"] == "equity_swap"
    assert by_ticker["IRS-1"] == "interest_rate_swap"
    assert by_ticker["CPI-1"] == "inflation_swap"
    assert by_ticker["MYSTERY-1"] == "other"  # unmapped asset type falls back, not dropped


def test_non_holdings_sheet_is_skipped_not_misparsed(tmp_path):
    wb = openpyxl.Workbook()
    _build_cover_sheet(wb, institution_reg_number="1")
    _build_stocks_sheet(wb, [("1", "1", "X", "1", "X Ord", "IL1", "", "", "ILS", 0.5, 10)])
    ws = wb.create_sheet("סכום נכסים")  # explicitly-skipped summary sheet
    ws.append(["Category", "Value"])
    ws.append(["Cash", 100])
    path = _save(wb, tmp_path)

    funds = [{"id": 1, "institution_reg_number": "1", "track_number": "1"}]
    result = parse_holdings_filing(path, funds)
    assert "סכום נכסים" not in result["sheets_parsed"]
    assert len(result["rows"]) == 1


def test_pct_of_track_is_read_as_a_fraction(tmp_path):
    wb = openpyxl.Workbook()
    _build_cover_sheet(wb, institution_reg_number="1")
    _build_stocks_sheet(wb, [("1", "1", "X", "1", "X Ord", "IL1", "", "", "ILS", 0.036151, 10)])
    path = _save(wb, tmp_path)

    funds = [{"id": 1, "institution_reg_number": "1", "track_number": "1"}]
    result = parse_holdings_filing(path, funds)
    assert result["rows"][0]["pct_of_track"] == pytest.approx(0.036151)


def test_header_matching_tolerates_trailing_whitespace(tmp_path):
    wb = openpyxl.Workbook()
    _build_cover_sheet(wb, institution_reg_number="1")
    ws = wb.create_sheet("מניות מבכ ויהש")
    # Trailing space + a Hebrew RTL mark, both confirmed to show up in real
    # regulatory spreadsheet headers.
    ws.append(["מספר מסלול ", "שם נייר ערך‏", "שיעור מנכסי אפיק ההשקעה", "שווי הוגן (באלפי ש\"ח)"])
    ws.append(["1", "X Ord", 0.4, 10])
    path = _save(wb, tmp_path)

    funds = [{"id": 1, "institution_reg_number": "1", "track_number": "1"}]
    result = parse_holdings_filing(path, funds)
    assert len(result["rows"]) == 1
    assert result["rows"][0]["security_name"] == "X Ord"


def test_zero_weight_placeholder_row_is_skipped(tmp_path):
    """Confirmed real-world case: some sheets emit an explicit zero row for
    a track that holds nothing in that category (kupa+track populated,
    every other cell blank) rather than omitting the track."""
    wb = openpyxl.Workbook()
    _build_cover_sheet(wb, institution_reg_number="1")
    _build_stocks_sheet(wb, [
        ("1", "1", "", "", "", "", "", "", "", 0, 0),
        ("1", "1", "Real Co", "1", "Real Ord", "IL1", "", "", "ILS", 0.5, 10),
    ])
    path = _save(wb, tmp_path)

    funds = [{"id": 1, "institution_reg_number": "1", "track_number": "1"}]
    result = parse_holdings_filing(path, funds)
    assert len(result["rows"]) == 1
    assert result["rows"][0]["security_name"] == "Real Ord"


def test_shared_institution_track_pair_matches_every_fund(tmp_path):
    """Real case (Altshuler Shaham study funds, confirmed with the user's
    own insurance agent): multiple of the user's own accounts can be pooled
    into the SAME investment track at one company. A matching row must be
    attributed to EVERY fund sharing that (institution, track) — not just
    whichever fund happens to be last in the lookup — so each fund's own
    weight calculation downstream sees the full track composition."""
    wb = openpyxl.Workbook()
    _build_cover_sheet(wb, institution_reg_number="1")
    _build_stocks_sheet(wb, [("1", "1", "X", "1", "X Ord", "IL1", "", "", "ILS", 0.5, 10)])
    path = _save(wb, tmp_path)

    funds = [
        {"id": 1, "institution_reg_number": "1", "track_number": "1"},
        {"id": 2, "institution_reg_number": "1", "track_number": "1"},
    ]
    result = parse_holdings_filing(path, funds)
    assert len(result["rows"]) == 2
    assert {r["fund_id"] for r in result["rows"]} == {1, 2}
    assert result["rows"][0]["security_number"] == result["rows"][1]["security_number"] == "IL1"


def test_unmatched_institution_produces_no_rows(tmp_path):
    wb = openpyxl.Workbook()
    _build_cover_sheet(wb, institution_reg_number="1")
    _build_stocks_sheet(wb, [("1", "1", "X", "1", "X Ord", "IL1", "", "", "ILS", 0.5, 10)])
    path = _save(wb, tmp_path)

    # No fund matches this institution at all — the sheet still resolves
    # structurally (so this is NOT the "no holdings sheets recognized" case,
    # which raises), it just contributes zero matched rows.
    funds = [{"id": 1, "institution_reg_number": "999", "track_number": "1"}]
    result = parse_holdings_filing(path, funds)
    assert result["rows"] == []
    assert result["unmatched_track_count"] == 1
