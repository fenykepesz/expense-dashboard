"""
Look-Through Holdings Filing Parser.

Parses an Israeli institutional "מצבת נכסים" (Uniform Structure asset
statement) — a quarterly regulatory filing, one .xlsx per institution, with
~30 sheets: a cover sheet, a summary sheet, a column-legend sheet, and ~20
"holdings" sheets (one per instrument type: cash, government bonds,
corporate bonds, traded/non-traded stocks, ETFs, mutual funds, warrants,
options, futures, structured products, loans, deposits, real estate, etc).

Every holdings sheet carries the same core fields by HEADER TEXT, not fixed
column position — sheets differ in column count (confirmed range 16-53) but
not in what the columns mean. Rows are filtered to the caller's own funds'
(institution_reg_number, track_number) pairs at parse time — an
institution's full filing can span dozens of unrelated tracks and tens of
thousands of rows; only the caller's matched rows are ever materialized.

Usage (standalone):
    python holdings_filing_to_json.py filing.xlsx --funds funds.json
"""

import argparse
import json
import re
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))

import openpyxl


# ── Cover sheet ───────────────────────────────────────────────────────────────

COVER_SHEET_NAME = "עמוד פתיחה"
SUMMARY_SHEET_NAME = "סכום נכסים"
LEGEND_SHEET_NAME = "מיפוי סעיפים"

COVER_LABELS = {
    "institution_reg_number": ["ח.פ. הגוף המוסדי"],
    "institution_name": ["הגוף המוסדי"],  # checked AFTER the ח.פ. variant, see _extract_cover_metadata
    "period_quarter": ["רבעון הדיווח"],
    "period_year": ["שנת הדיווח"],
}


def _normalize_header(text):
    """Strip RTL/LTR direction marks and collapse whitespace — real
    spreadsheets routinely have trailing spaces or invisible marks in
    Hebrew headers."""
    if text is None:
        return ""
    text = str(text).replace("‎", "").replace("‏", "")
    return re.sub(r"\s+", " ", text).strip()


def _extract_cover_metadata(ws):
    """Label-scan the cover sheet: for each row, treat every non-blank cell
    except the last as "label text" and the last non-blank cell as the
    value. Deliberately doesn't hardcode which column holds which —
    confirmed to shift by institution (one real sample's cover sheet has
    the label in column A; another has an extra leading marker column,
    shifting the label to column B). Filtering to non-blank cells only
    and taking "everything before the last one" as label context survives
    that shift without caring which absolute column either lands in."""
    found = {}
    for row in ws.iter_rows(min_row=1, max_row=30, values_only=True):
        non_blank = [cell for cell in row if cell is not None and _normalize_header(cell) != ""]
        if len(non_blank) < 2:
            continue
        value = non_blank[-1]
        label_text = " ".join(_normalize_header(c) for c in non_blank[:-1])
        if "institution_reg_number" not in found and any(l in label_text for l in COVER_LABELS["institution_reg_number"]):
            found["institution_reg_number"] = str(value).strip()
        elif "institution_name" not in found and any(l in label_text for l in COVER_LABELS["institution_name"]):
            found["institution_name"] = str(value).strip()
        elif "period_quarter" not in found and any(l in label_text for l in COVER_LABELS["period_quarter"]):
            found["period_quarter"] = value
        elif "period_year" not in found and any(l in label_text for l in COVER_LABELS["period_year"]):
            found["period_year"] = value
    return found


def _to_int(value):
    try:
        return int(str(value).strip())
    except (TypeError, ValueError):
        return None


# ── Holdings sheets ───────────────────────────────────────────────────────────

# {internal field name: [known Hebrew header text variants]}. Matched by
# normalized header TEXT against whichever sheet's own header row, not by
# position — this is what survives the 16-53 column-count variance across
# sheet types. A sheet missing an optional field just yields '' for it.
CANONICAL_FIELDS = {
    "kupa_number":     ["מספר קופה/קרן/ח.פ.", "מספר קופה/קרן/ח.פ. עבור חברת ביטוח", "מספר קופה"],
    "track_number":    ["מספר מסלול"],
    "issuer_name":     ["שם מנפיק", "שם הבנק", "צד נגדי - Counterparty"],
    "issuer_number":   ["מספר מנפיק", "מספר מזהה בנק", "מספר מזהה לווה"],
    # Non-securities (cash/deposits/loans/real estate/investment funds/
    # derivatives/other) don't have a "security" in the usual sense — each
    # sheet names its own holding differently (bank, loan, property, fund,
    # ticker...) but always exactly one of these, so there's no ambiguity
    # in which variant a given sheet actually uses.
    "security_name":   ["שם נייר ערך", "שם הלוואה", "שם הנכס", "שם הנכס האחר",
                         "שם קרן השקעה", "טיקר"],
    "security_number": ["מספר נייר ערך", "מספר הלוואה", "מספר הנכס האחר",
                         "מספר מזהה קרן השקעה"],
    "country":         ["מדינה לפי חשיפה כלכלית"],
    "sector":          ["ענף מסחר"],
    "currency":        ["מטבע פעילות"],
    # NOTE: despite its label ("% of track's assets"), this column is
    # confirmed (via real Phoenix file, cross-checked against two bond
    # sheets and the cash sheet) to sum to ~100% WITHIN EACH INSTRUMENT-TYPE
    # SHEET for a given track, not across the track's total value — one
    # sheet had a row at 525%. It is NOT usable for dollar math and is kept
    # purely as an informational per-category weight.
    "pct_of_track":    ["שיעור מנכסי אפיק ההשקעה"],
    # Absolute fair value in thousands ILS — confirmed present on every real
    # holdings sheet type and, unlike pct_of_track, already scoped to the
    # specific track (summing it across all of one track's sheets produced
    # a sane total that matched the track's real order of magnitude). This
    # is the actual dollar-value source; converted to real ILS (×1000) at
    # parse time.
    "fair_value_ils":  ["שווי הוגן (באלפי ש\"ח)", "שווי הוגן (נטו באלפי ש\"ח)"],
    # Only meaningful on the OTHER_DERIVATIVES_SHEET_NAME sheet (see below) —
    # that one sheet bundles several genuinely different derivative types
    # (FX, interest rate, equity, inflation) under one sheet name, and this
    # is the filing's own column distinguishing which is which per row.
    "asset_type":      ["סוג הנכס"],
}

# Fields that must all resolve for a sheet to be treated as holdings-shaped.
# fair_value_ils drives the dollar math; pct_of_track is informational only
# (see note above) and is NOT required.
REQUIRED_FIELDS = {"track_number", "fair_value_ils"}

# Confirmed real sheet names -> instrument type. A holdings-shaped sheet
# whose name isn't here still imports (tagged 'other') and is recorded as a
# warning — formats drift between companies/years, degrade gracefully
# rather than reject the whole file.
SHEET_NAME_TO_INSTRUMENT_TYPE = {
    "מזומנים ושווי מזומנים": "cash",
    "איגרות חוב ממשלתיות": "govt_bond",
    "ניירות ערך מסחריים": "corp_bond",
    "איגרות חוב": "corp_bond",
    "מניות מבכ ויהש": "equity_traded",
    "קרנות סל": "etf",
    "קרנות נאמנות": "mutual_fund",
    "כתבי אופציה": "warrant",
    "אופציות": "option",
    "חוזים עתידיים": "future",
    "מוצרים מובנים": "structured_product",
    "לא סחיר איגרות חוב ממשלתיות": "govt_bond",
    "לא סחיר איגרות חוב מיועדות": "govt_bond",
    "אפיק השקעה מובטח תשואה": "structured_product",
    "לא סחיר ניירות ערך מסחריים": "corp_bond",
    "לא סחיר איגרות חוב": "corp_bond",
    "לא סחיר מניות מבכ ויהש": "equity_nontraded",
    "קרנות השקעה": "investment_fund",
    "לא סחיר כתבי אופציה": "warrant",
    "לא סחיר אופציות": "option",
    "לא סחיר נגזרים אחרים": "other",  # overridden per-row below, see OTHER_DERIVATIVES_SHEET_NAME
    "הלוואות": "loan",
    "לא סחיר מוצרים מובנים": "structured_product",
    "פיקדונות מעל 3 חודשים": "deposit",
    "זכויות מקרקעין": "real_estate",
    "השקעה בחברות מוחזקות": "investment_fund",
    "נכסים אחרים": "other",
}

SKIPPED_SHEETS = {COVER_SHEET_NAME, SUMMARY_SHEET_NAME, LEGEND_SHEET_NAME, "File Name Info",
                  "אפשרויות בחירה", "מסגרות אשראי", "יתרות התחייבות להשקעה"}

# This one sheet genuinely bundles several different OTC derivative types
# under a single sheet name (confirmed via real data: 5,970 FX rows, 714
# interest-rate rows, 440 equity/index rows, 36 inflation rows, all in the
# same sheet) — a sheet-name-level mapping can't tell them apart, but the
# sheet carries its own "סוג הנכס" (Asset Type) column that reliably can.
# This is the one place classification reads row content rather than just
# sheet/header names — deliberately narrow (one sheet, one filing-defined
# column with a small closed set of real values), not a general heuristic.
OTHER_DERIVATIVES_SHEET_NAME = "לא סחיר נגזרים אחרים"
ASSET_TYPE_TO_INSTRUMENT_TYPE = {
    'מט"ח': "fx_swap",
    "ריבית ואג\"ח": "interest_rate_swap",
    "מניות לרבות מדדי מניות": "equity_swap",
    "מדד המחירים לצרכן": "inflation_swap",
}


def _resolve_header(header_row):
    """Given a tuple of header cell values, return {internal_name: col_index}
    for whichever CANONICAL_FIELDS entries match."""
    normalized = [_normalize_header(c) for c in header_row]
    resolved = {}
    for internal_name, variants in CANONICAL_FIELDS.items():
        for i, cell_text in enumerate(normalized):
            if cell_text and any(cell_text == v or cell_text.startswith(v) for v in variants):
                resolved[internal_name] = i
                break
    return resolved


def _find_header_row(ws):
    """Holdings sheets have their header on row 1 in every confirmed sample,
    but scan the first 3 rows and pick whichever resolves the most
    CANONICAL_FIELDS, in case a sheet has a title row above the real
    header."""
    best_row_idx, best_resolved, best_score = None, {}, -1
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=3, values_only=True), start=1):
        resolved = _resolve_header(row)
        if len(resolved) > best_score:
            best_row_idx, best_resolved, best_score = row_idx, resolved, len(resolved)
    return best_row_idx, best_resolved


def _to_float(cell):
    if cell is None:
        return 0.0
    if isinstance(cell, (int, float)):
        return float(cell)
    try:
        return float(str(cell).replace(",", "").strip())
    except ValueError:
        return 0.0


def parse_holdings_filing(path, funds):
    """Parse one institution's uniform-structure filing.

    `funds` is the full db.get_funds()-shaped list — used to build the
    (institution_reg_number, track_number) -> fund_id lookup that scopes
    every holdings sheet down to only the caller's own matched rows.

    Returns:
        {
          'institution_name', 'institution_reg_number',
          'period_year', 'period_quarter',
          'rows': [ {fund_id, instrument_type, issuer_name, issuer_number,
                     security_name, security_number, pct_of_track,
                     fair_value_ils, country, sector, currency}, ... ],
          'unmatched_track_count': int,
          'unrecognized_sheets': [ {'sheet_name', 'reason'}, ... ],
          'sheets_parsed': [str, ...],
        }

    Raises ValueError if the file doesn't look like a מצבת נכסים filing at
    all (no cover sheet, zero holdings sheets recognized, or the file isn't
    even a valid .xlsx — never lets openpyxl's lower-level exceptions
    escape as an unhandled 500).
    """
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        raise ValueError(f"Couldn't open this file as an Excel (.xlsx) workbook: {e}")
    try:
        cover_name = next((n for n in wb.sheetnames if COVER_SHEET_NAME in n), None)
        if cover_name is None:
            raise ValueError(
                "This doesn't look like a מצבת נכסים (Uniform Structure) filing — "
                f'no "{COVER_SHEET_NAME}" cover sheet found.'
            )
        cover = _extract_cover_metadata(wb[cover_name])
        institution_reg_number = cover.get("institution_reg_number", "")
        institution_name = cover.get("institution_name", "")
        period_year = _to_int(cover.get("period_year"))
        period_quarter = _to_int(cover.get("period_quarter"))
        if not institution_reg_number or period_year is None or period_quarter is None:
            raise ValueError(
                "Couldn't read the institution registration number and filing period "
                "from the cover sheet — the file's layout may not match the expected format."
            )

        # (institution_reg_number, track_number) -> fund_id, scoped to funds
        # actually belonging to this institution.
        track_lookup = {
            (f["institution_reg_number"], f["track_number"]): f["id"]
            for f in funds
            if f.get("institution_reg_number") == institution_reg_number and f.get("track_number")
        }

        rows = []
        unrecognized_sheets = []
        sheets_parsed = []
        unmatched_track_count = 0

        for sheet_name in wb.sheetnames:
            if sheet_name in SKIPPED_SHEETS or COVER_SHEET_NAME in sheet_name:
                continue
            ws = wb[sheet_name]
            header_row_idx, resolved = _find_header_row(ws)
            if not REQUIRED_FIELDS.issubset(resolved) or (
                "security_name" not in resolved and "issuer_name" not in resolved
            ):
                unrecognized_sheets.append({
                    "sheet_name": sheet_name,
                    "reason": "Doesn't look like a holdings sheet (missing track/%/security columns) — skipped.",
                })
                continue
            # `resolved` being a superset of REQUIRED_FIELDS (just checked above)
            # is only possible when _find_header_row actually matched a row.
            assert header_row_idx is not None

            instrument_type = SHEET_NAME_TO_INSTRUMENT_TYPE.get(sheet_name)
            if instrument_type is None:
                instrument_type = "other"
                unrecognized_sheets.append({
                    "sheet_name": sheet_name,
                    "reason": "Holdings-shaped sheet with an unrecognized name — "
                              "imported as instrument_type='other'.",
                })

            sheets_parsed.append(sheet_name)

            def get(row, field, default=""):
                idx = resolved.get(field)
                return row[idx] if idx is not None and idx < len(row) else default

            for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
                track_number = str(get(row, "track_number") or "").strip()
                fund_id = track_lookup.get((institution_reg_number, track_number))
                if fund_id is None:
                    if track_number:
                        unmatched_track_count += 1
                    continue

                issuer_name = str(get(row, "issuer_name") or "").strip()
                issuer_number = str(get(row, "issuer_number") or "").strip()
                security_name = str(get(row, "security_name") or "").strip()
                security_number = str(get(row, "security_number") or "").strip()
                pct_of_track = _to_float(get(row, "pct_of_track", 0))
                # Source values are in thousands ILS; store real ILS.
                fair_value_ils = _to_float(get(row, "fair_value_ils", 0)) * 1000
                # Some sheets emit an explicit zero row for a track that
                # simply holds nothing in that category (confirmed on the
                # investment-funds sheet: kupa+track populated, every other
                # cell blank) rather than omitting the track. Zero value
                # AND no identifying field at all carries no information —
                # skip it rather than storing pure noise.
                if fair_value_ils == 0 and pct_of_track == 0 and not (
                    issuer_name or issuer_number or security_name or security_number
                ):
                    continue

                row_instrument_type = instrument_type
                if sheet_name == OTHER_DERIVATIVES_SHEET_NAME:
                    asset_type = str(get(row, "asset_type") or "").strip()
                    row_instrument_type = ASSET_TYPE_TO_INSTRUMENT_TYPE.get(asset_type, instrument_type)

                rows.append({
                    "fund_id": fund_id,
                    "instrument_type": row_instrument_type,
                    "issuer_name": issuer_name,
                    "issuer_number": issuer_number,
                    "security_name": security_name,
                    "security_number": security_number,
                    "pct_of_track": pct_of_track,
                    "fair_value_ils": fair_value_ils,
                    "country": str(get(row, "country") or "").strip(),
                    "sector": str(get(row, "sector") or "").strip(),
                    "currency": str(get(row, "currency") or "").strip(),
                })

        if not sheets_parsed:
            raise ValueError(
                "No holdings sheets were recognized in this file — it may not be a "
                "מצבת נכסים filing, or its structure has drifted from the expected format."
            )

        return {
            "institution_name": institution_name,
            "institution_reg_number": institution_reg_number,
            "period_year": period_year,
            "period_quarter": period_quarter,
            "rows": rows,
            "unmatched_track_count": unmatched_track_count,
            "unrecognized_sheets": unrecognized_sheets,
            "sheets_parsed": sheets_parsed,
        }
    finally:
        wb.close()


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Parse a look-through holdings filing (.xlsx)")
    parser.add_argument("filing", help="Path to the institution's uniform-structure .xlsx file")
    parser.add_argument("--funds", help="Path to a JSON file with a funds list (db.get_funds() shape)")
    parser.add_argument("-o", "--output", help="Write the parsed result as JSON to this path")
    args = parser.parse_args()

    funds_list = []
    if args.funds:
        with open(args.funds, encoding="utf-8") as f:
            funds_list = json.load(f)

    result = parse_holdings_filing(args.filing, funds_list)
    output_json = json.dumps(result, indent=2, ensure_ascii=False)
    if args.output:
        Path(args.output).write_text(output_json, encoding="utf-8")
        print(f"Wrote {len(result['rows'])} matched rows to {args.output}")
    else:
        print(output_json)
